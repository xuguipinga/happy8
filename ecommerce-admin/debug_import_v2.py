
import pandas as pd
from decimal import Decimal
import io
import os
import openpyxl

def test_openpyxl_parser():
    file_path = r'd:\WorkSpec\Project\Ecommerce-admin\ecommerce-admin\docs\克罗心手镯库存清单.xlsx'
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    for ws in wb.worksheets:
        name = ws.title
        max_row = ws.max_row
        max_col = ws.max_column
        current_series = None
        last_model_in_col = {}

        r = 1
        group_size = 0
        has_spec = False
        
        while r <= max_row:
            row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, max_col + 1)]
            if not any(row_vals): 
                r += 1
                continue

            v1 = row_vals[0]
            if '系列' in v1 and ('清单' in v1 or '库存' in v1 or '系列' in v1):
                current_series = v1
                last_model_in_col = {}
                r += 1
                continue

            if '型号' in row_vals:
                idx_model = row_vals.index('型号')
                if idx_model + 2 < len(row_vals) and '数量' in row_vals[idx_model + 1:idx_model + 3]:
                    if row_vals[idx_model + 2] == '数量':
                        group_size = 3
                        has_spec = True
                    else:
                        group_size = 2
                        has_spec = False
                elif idx_model + 1 < len(row_vals) and row_vals[idx_model + 1] == '数量':
                    group_size = 2
                    has_spec = False
                
                if group_size > 0:
                    r += 1
                    continue

            if group_size > 0:
                for c_start in range(1, max_col + 1, group_size):
                    if c_start > max_col: break
                    model = str(ws.cell(r, c_start).value or '').strip()
                    
                    if model and model.lower() not in ['none', 'nan']:
                        last_model_in_col[c_start] = model
                    else:
                        model = last_model_in_col.get(c_start, '')
                    
                    spec = ""
                    qty_val = None
                    
                    if group_size == 3:
                        spec = str(ws.cell(r, c_start + 1).value or '').strip()
                        qty_val = ws.cell(r, c_start + 2).value
                    else:
                        qty_val = ws.cell(r, c_start + 1).value
                    
                    if not model:
                        continue
                        
                    print(f"Row: {r}, Col: {c_start}, Model: '{model}', Spec: '{spec}', Qty_val: '{qty_val}' (type: {type(qty_val)})")
                r += 1
                continue
            r += 1

if __name__ == '__main__':
    test_openpyxl_parser()

def debug_excel_file(file_path):
    print(f"--- Debugging File Struct: {file_path} ---")
    if not os.path.exists(file_path):
        print("Error: File not found!")
        return

    try:
        df_raw = pd.read_excel(file_path, header=None)
        print("\n[Row 0-5 Snapshot]:")
        for i in range(min(10, len(df_raw))):
            print(f"Row {i}: {list(df_raw.iloc[i])}")
            
        with open(file_path, 'rb') as f:
            content = f.read()
        
        results = parse_inventory_excel(content)
        
        print(f"\n[Extraction results]: {len(results)} items found.")
        if results:
            for item in results[:10]:
                print(item)
        else:
            print("FAILED: No items found.")

    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    target = r"d:\WorkSpec\Project\Ecommerce-admin\ecommerce-admin\docs\克罗心手镯库存清单.xlsx"
    debug_excel_file(target)
