import pandas as pd
from decimal import Decimal
import io
import os
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
from app.utils.image_helper import save_image_from_bytes

def extract_images_from_xlsx(file_content):
    """
    解析 XLSX 中的所有图片，支持多 Sheet。
    返回: {(sheet_name, row, col): image_bytes}
    """
    image_map = {}
    try:
        with zipfile.ZipFile(io.BytesIO(file_content), 'r') as z:
            # 1. 先建立 rid -> image_path 的全局映射 (从 xl/_rels/workbook.xml.rels 里可能找不到，直接从 xl/drawings/_rels 下找所有的)
            # 实际上每个 sheet 有自己的 drawing，我们需要遍历所有的 sheet
            
            # 读取 workbook.xml 获取 sheetId 与 name 的对应关系
            workbook_xml = z.read('xl/workbook.xml')
            wb_root = ET.fromstring(workbook_xml)
            # 命名空间处理
            ns_main = {'ns': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main', 
                       'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheets_info = {} # rId -> name
            for s in wb_root.findall('.//ns:sheet', ns_main):
                name = s.attrib.get('name')
                rid = s.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                sheets_info[rid] = name

            # 读取 xl/_rels/workbook.xml.rels 获取 rId -> target (path to sheet.xml)
            wb_rels_xml = z.read('xl/_rels/workbook.xml.rels')
            wb_rels_root = ET.fromstring(wb_rels_xml)
            ns_rel = {'r': 'http://schemas.openxmlformats.org/package/2006/relationships'}
            
            sheet_path_map = {} # name -> sheet_xml_path (e.g. xl/worksheets/sheet1.xml)
            for r in wb_rels_root.findall('.//r:Relationship', ns_rel):
                rid = r.attrib.get('Id')
                if rid in sheets_info:
                    target = r.attrib.get('Target')
                    sheet_path_map[sheets_info[rid]] = f"xl/{target}"

            # 遍历每个 sheet 找对应的 drawing
            for sheet_name, sheet_xml_path in sheet_path_map.items():
                rels_path = f"xl/worksheets/_rels/{os.path.basename(sheet_xml_path)}.rels"
                if rels_path not in z.namelist():
                    continue
                
                rels_root = ET.fromstring(z.read(rels_path))
                drawing_path = None
                for rel in rels_root:
                    if 'drawing' in rel.attrib.get('Type', ''):
                        target = rel.attrib.get('Target')
                        drawing_path = target.replace('../drawings/', 'xl/drawings/')
                        break
                
                if not drawing_path: continue
                
                # 读取 drawing 的 rels
                drawing_rels_path = f'xl/drawings/_rels/{os.path.basename(drawing_path)}.rels'
                media_map = {}
                if drawing_rels_path in z.namelist():
                    draw_rels_root = ET.fromstring(z.read(drawing_rels_path))
                    for r in draw_rels_root:
                        rid = r.attrib.get('Id')
                        target = r.attrib.get('Target').replace('../media/', 'xl/media/')
                        media_map[rid] = target
                
                # 解析 drawing 获取坐标
                draw_root = ET.fromstring(z.read(drawing_path))
                ns_draw = {
                    'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
                    'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
                    'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
                }
                
                for anchor in draw_root.findall('.//xdr:twoCellAnchor', ns_draw) + draw_root.findall('.//xdr:oneCellAnchor', ns_draw):
                    from_el = anchor.find('.//xdr:from', ns_draw)
                    if from_el is None: continue
                    row = int(from_el.find('xdr:row', ns_draw).text)
                    col = int(from_el.find('xdr:col', ns_draw).text)
                    
                    pic = anchor.find('.//xdr:pic', ns_draw)
                    if pic is None: continue
                    blip = pic.find('.//a:blip', ns_draw)
                    if blip is None: continue
                    
                    rid = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                    if rid in media_map:
                        image_map[(sheet_name, row, col)] = z.read(media_map[rid])
                        
    except Exception as e:
        print(f"Global image extraction error: {e}")
        
    return image_map

def parse_inventory_excel(file_content):
    """
    解析库存 Excel，支持多 Sheet 和多种混合格式。
    """
    # 提取图片: {(sheet_name, row, col): image_bytes}
    image_map = extract_images_from_xlsx(file_content)
    
    # 使用 openpyxl 读取数据
    wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
    results = []
    
    for ws in wb.worksheets:
        name = ws.title
        max_row = ws.max_row
        max_col = ws.max_column
        current_series = None
        
        # 记录每组列的上一个型号，用于处理合并单元格/追溯型号
        # key: col_index, value: last_model_name
        last_model_in_col = {}

        r = 1
        group_size = 0
        has_spec = False

        while r <= max_row:
            row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, max_col + 1)]
            if not any(row_vals): 
                r += 1
                continue

            v1 = row_vals[0]
            
            # 1. 检测系列行 (如 "克罗心库存清单C系列")
            if '系列' in v1 and ('清单' in v1 or '库存' in v1 or '系列' in v1):
                current_series = v1
                # 切换系列时，重置型号追溯
                last_model_in_col = {}
                r += 1
                continue

            # 2. 识别表头并确定模式
            # 检查当前行是否包含表头关键字
            if '型号' in row_vals:
                # ==== 新增：标准纵向表格解析模式 ====
                if row_vals.count('型号') == 1 and ('数量' in row_vals or '现有库存' in row_vals):
                    col_map = {name: idx for idx, name in enumerate(row_vals)}
                    r += 1
                    while r <= max_row:
                        curr_row_vals = [str(ws.cell(r, c).value or '').strip() for c in range(1, max_col + 1)]
                        if not any(curr_row_vals):
                            r += 1
                            continue
                        
                        model = curr_row_vals[col_map['型号']] if '型号' in col_map and col_map['型号'] < len(curr_row_vals) else ''
                        if not model:
                            r += 1
                            continue
                            
                        spec = curr_row_vals[col_map['规格']] if '规格' in col_map and col_map['规格'] < len(curr_row_vals) else ''
                        
                        qty_col = col_map.get('数量') if '数量' in col_map else col_map.get('现有库存')
                        qty_val = curr_row_vals[qty_col] if qty_col is not None and qty_col < len(curr_row_vals) else '0'
                        try: qty = Decimal(qty_val)
                        except: qty = Decimal('0')
                        
                        cost_col = col_map.get('平均成本')
                        cost_val = curr_row_vals[cost_col] if cost_col is not None and cost_col < len(curr_row_vals) else '0'
                        try: avg_cost = Decimal(cost_val)
                        except: avg_cost = Decimal('0')
                        
                        series_col = col_map.get('系列') if '系列' in col_map else col_map.get('产品系列')
                        series = curr_row_vals[series_col] if series_col is not None and series_col < len(curr_row_vals) else current_series
                        if not series and model and model[0].isalpha():
                            series = model[0].upper() + '系列'
                        
                        img_col = col_map.get('图片') if '图片' in col_map else col_map.get('示例图')
                        img_url = None
                        if img_col is not None:
                            for anchor in [(r-1, img_col)]:
                                if (name, anchor[0], anchor[1]) in image_map:
                                    img_url = save_image_from_bytes(image_map[(name, anchor[0], anchor[1])], f"{model}.jpg")
                                    break
                        else:
                            for anchor in [(r-1, 0), (r-2, 0), (r-3, 0)]:
                                if (name, anchor[0], anchor[1]) in image_map:
                                    img_url = save_image_from_bytes(image_map[(name, anchor[0], anchor[1])], f"{model}.jpg")
                                    break
                                    
                        results.append({
                            'model': model,
                            'spec': spec,
                            'quantity': qty,
                            'avg_cost': avg_cost,
                            'image_url': img_url,
                            'series': series
                        })
                        r += 1
                    continue # 当前 sheet 按标准表解析完毕，跳出当前组逻辑
                
                # 探测组大小
                idx_model = row_vals.index('型号')
                if idx_model + 2 < len(row_vals) and '数量' in row_vals[idx_model + 1:idx_model + 3]:
                    if row_vals[idx_model + 2] == '数量':
                        group_size = 3
                        has_spec = True
                    else:
                        group_size = 2
                        has_spec = False
                elif idx_model + 1 < len(row_vals) and row_vals[idx_model + 1] == '数量':
                    group_size = 2
                    has_spec = False
                
                if group_size > 0:
                    r += 1 # 跳过表头行
                    continue

            # 3. 如果已经确定了组大小，且当前行不是系列行也不是空行，则尝试按组解析
            if group_size > 0:
                has_any_data = False
                # 按组横向遍历
                for c_start in range(1, max_col + 1, group_size):
                    if c_start > max_col: break
                    model = str(ws.cell(r, c_start).value or '').strip()
                    
                    # 获取规格和数量
                    spec = ""
                    qty_val = None
                    
                    if group_size == 3:
                        spec = str(ws.cell(r, c_start + 1).value or '').strip()
                        qty_val = ws.cell(r, c_start + 2).value
                    else:
                        qty_val = ws.cell(r, c_start + 1).value
                    
                    # 处理型号追溯 (Carry-over)
                    if model and model.lower() not in ['none', 'nan']:
                        last_model_in_col[c_start] = model
                    else:
                        model = last_model_in_col.get(c_start, '')
                    
                    if not model: continue
                    
                    # 获取规格和数量
                    spec = ""
                    qty_val = None
                    
                    if group_size == 3:
                        spec = str(ws.cell(r, c_start + 1).value or '').strip()
                        qty_val = ws.cell(r, c_start + 2).value
                    else:
                        qty_val = ws.cell(r, c_start + 1).value
                    
                    # 检查这一组是否有任何形式的数据指示 (型号已在上面判断过)
                    # 只要有型号，我们就允许这行数据通过，即使数量为空，也会被视为 0
                    if not (str(ws.cell(r, c_start).value or '').strip()) and not spec:
                        continue

                    has_any_data = True
                    qty = Decimal('0')
                    try:
                        if qty_val is not None and str(qty_val).strip() != '':
                            qty = Decimal(str(qty_val))
                    except: pass
                    
                    # 图片提取
                    img_url = None
                    for anchor in [(r-1, c_start-1), (r-2, c_start-1), (r-3, c_start-1)]:
                        if (name, anchor[0], anchor[1]) in image_map:
                            img_url = save_image_from_bytes(image_map[(name, anchor[0], anchor[1])], f"{model}.jpg")
                            break

                    results.append({
                        'model': model,
                        'spec': spec,
                        'quantity': qty,
                        'avg_cost': Decimal('0'),
                        'image_url': img_url,
                        'series': current_series if current_series else (model[0].upper() + '系列' if model and model[0].isalpha() else None)
                    })
                r += 1
                continue
            
            # 4. 兜底逻辑：传统的 3 行网格模式
            if v1 in ['编号', '型号'] and r + 2 <= max_row:
                 for c in range(2, max_col + 1):
                    model = str(ws.cell(r, c).value or '').strip()
                    if not model or model.lower() in ['none', 'nan']: continue
                    
                    spec = str(ws.cell(r + 1, c).value or '').strip()
                    price_val = ws.cell(r + 2, c).value
                    avg_cost = Decimal('0')
                    try:
                        if price_val is not None: avg_cost = Decimal(str(price_val))
                    except: pass
                    
                    img_url = None
                    for anchor in [(r-1, c-1), (r-2, c-1), (r-3, c-1)]:
                        if (name, anchor[0], anchor[1]) in image_map:
                            img_url = save_image_from_bytes(image_map[(name, anchor[0], anchor[1])], f"{model}.jpg")
                            break

                    results.append({
                        'model': model,
                        'spec': spec,
                        'quantity': Decimal('0'),
                        'avg_cost': avg_cost,
                        'image_url': img_url,
                        'series': current_series if current_series else (model[0].upper() + '系列' if model and model[0].isalpha() else None)
                    })
                 r += 3
                 continue

            r += 1
            
    return results
