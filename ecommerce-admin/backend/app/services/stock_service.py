from decimal import Decimal
from app.extensions import db
from app.models.stock import Inventory, StockRecord
from app.repositories.inventory_repository import InventoryRepository
from app.repositories.stock_record_repository import StockRecordRepository

def to_decimal(value, default=Decimal('0')):
    if value is None or (isinstance(value, str) and value.strip() == ''):
        return default
    try:
        return Decimal(str(value))
    except:
        return default


import io
import os
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image as OpenpyxlImage
from flask import current_app
from datetime import datetime

class StockService:
    @classmethod
    def get_inventory_list(cls, tenant_id, search=None, status=None, series=None, page=1, per_page=20, sort_by='model', sort_order='ascending'):
        query = InventoryRepository.find_by_tenant(tenant_id, search, status, series)
        
        # 排序逻辑
        if sort_by == 'model':
            if sort_order == 'ascending':
                query = query.order_by(Inventory.model.asc(), Inventory.spec.asc())
            else:
                query = query.order_by(Inventory.model.desc(), Inventory.spec.desc())
        elif sort_by == 'updated_at':
            query = query.order_by(Inventory.updated_at.asc() if sort_order == 'ascending' else Inventory.updated_at.desc())
        
        return query.paginate(page=page, per_page=per_page)

    @classmethod
    def create_inventory(cls, tenant_id, data, operator_name):
        model = data.get('model')
        spec = data.get('spec', '')
        
        if not model:
            raise ValueError("型号不能为空")
            
        exists = InventoryRepository.find_by_model_and_spec(tenant_id, model, spec)
        if exists:
            raise ValueError("该型号规格已存在")
            
        initial_qty = to_decimal(data.get('quantity', 0))
        avg_cost = to_decimal(data.get('avg_cost', 0))

        
        inventory = Inventory(
            tenant_id=tenant_id,
            model=model,
            spec=spec,
            status=data.get('status', 'NORMAL'),
            quantity=initial_qty,
            unit=data.get('unit', 'pcs'),
            avg_cost=avg_cost,
            image_url=data.get('image_url'),
            series=data.get('series')
        )
        
        db.session.add(inventory)
        db.session.flush()
        
        if initial_qty != 0:
            record = StockRecord(
                tenant_id=tenant_id,
                inventory_id=inventory.id,
                record_type='IN' if initial_qty > 0 else 'OUT',
                change_quantity=initial_qty,
                balance_quantity=initial_qty,
                unit_cost=avg_cost,
                remark='初始化库存',
                operator_name=operator_name
            )
            db.session.add(record)
            
        db.session.commit()
        return inventory

    @classmethod
    def adjust_inventory(cls, tenant_id, inventory_id, change_qty, record_type, operator_name, unit_cost=None, remark=''):
        inventory = InventoryRepository.get_by_id(inventory_id)
        if not inventory or inventory.tenant_id != tenant_id:
            raise ValueError("库存项目未找到")
            
        change_qty = to_decimal(change_qty)
        if record_type == 'IN' and unit_cost is not None and change_qty > 0:
            unit_cost = to_decimal(unit_cost)

            total_value = (inventory.quantity * inventory.avg_cost) + (change_qty * unit_cost)
            new_total_qty = inventory.quantity + change_qty
            if new_total_qty > 0:
                inventory.avg_cost = total_value / new_total_qty
        
        inventory.quantity += change_qty
        
        record = StockRecord(
            tenant_id=tenant_id,
            inventory_id=inventory.id,
            record_type=record_type,
            change_quantity=change_qty,
            balance_quantity=inventory.quantity,
            unit_cost=unit_cost or inventory.avg_cost,
            remark=remark,
            operator_name=operator_name
        )
        
        db.session.add(record)
        db.session.commit()
        return inventory

    @classmethod
    def get_stock_records(cls, tenant_id, inventory_id=None, page=1, per_page=20):
        query = StockRecordRepository.find_by_inventory(tenant_id, inventory_id)
        return query.paginate(page=page, per_page=per_page)

    @classmethod
    def update_inventory(cls, tenant_id, inventory_id, data, operator_name):
        inventory = InventoryRepository.get_by_id(inventory_id)
        if not inventory or inventory.tenant_id != tenant_id:
            raise ValueError("库存项目未找到")
            
        if 'model' in data: inventory.model = data['model']
        if 'spec' in data: inventory.spec = data['spec']
        if 'unit' in data: inventory.unit = data['unit']
        if 'avg_cost' in data: inventory.avg_cost = to_decimal(data['avg_cost'])

        if 'image_url' in data: inventory.image_url = data['image_url']
        if 'series' in data: inventory.series = data['series']
        
        if 'quantity' in data:
            new_qty = to_decimal(data['quantity'])

            if new_qty != inventory.quantity:
                change = new_qty - inventory.quantity
                record = StockRecord(
                    tenant_id=tenant_id,
                    inventory_id=inventory.id,
                    record_type='ADJ',
                    change_quantity=change,
                    balance_quantity=new_qty,
                    remark='手动修改数量',
                    operator_name=operator_name
                )
                db.session.add(record)
                inventory.quantity = new_qty
        
        db.session.commit()
        return inventory

    @classmethod
    def get_unique_series(cls, tenant_id):
        # 获取所有非空的系列
        series = db.session.query(Inventory.series).filter(
            Inventory.tenant_id == tenant_id,
            Inventory.series != None,
            Inventory.series != ''
        ).distinct().all()
        return [s[0] for s in series]

    @classmethod
    def delete_inventory(cls, tenant_id, inventory_id):
        inventory = InventoryRepository.get_by_id(inventory_id)
        if not inventory or inventory.tenant_id != tenant_id:
            raise ValueError("库存项目未找到")
            
        StockRecord.query.filter_by(inventory_id=inventory_id).delete()
        db.session.delete(inventory)
        db.session.commit()

    @classmethod
    def import_inventory(cls, tenant_id, file_content, clear_existing, operator_name, import_mode='all'):
        from app.utils.inventory_parser import parse_inventory_excel
        
        if clear_existing:
            inv_ids = [inv.id for inv in Inventory.query.filter_by(tenant_id=tenant_id).all()]
            if inv_ids:
                StockRecord.query.filter(StockRecord.inventory_id.in_(inv_ids)).delete(synchronize_session=False)
                Inventory.query.filter(Inventory.id.in_(inv_ids)).delete(synchronize_session=False)
                db.session.flush()

        items = parse_inventory_excel(file_content)
        count = 0
        new_models = 0
        
        for item in items:
            # 如果是只同步图片模式，只寻找已经存在的型号，不去匹配规格，也不创建新库存
            if import_mode == 'only_image':
                invs = Inventory.query.filter_by(
                    tenant_id=tenant_id, 
                    model=item['model']
                ).all()
                if not invs:
                    continue # 找不到对应的型号，直接跳过，不新增
            else:
                invs = Inventory.query.filter_by(
                    tenant_id=tenant_id, 
                    model=item['model'], 
                    spec=item['spec']
                ).all()
            
            if not invs:
                # 根据导入模式设置字段
                qty_to_set = item['quantity'] if import_mode in ['all', 'only_data'] else 0
                avg_cost_to_set = item.get('avg_cost', 0) if import_mode in ['all', 'only_data'] else 0
                image_url_to_set = item.get('image_url') if import_mode in ['all', 'only_image'] else None

                inv = Inventory(
                    tenant_id=tenant_id,
                    model=item['model'],
                    spec=item['spec'],
                    quantity=qty_to_set,
                    unit='pcs',
                    avg_cost=avg_cost_to_set,
                    image_url=image_url_to_set,
                    series=item.get('series')
                )
                db.session.add(inv)
                new_models += 1
                db.session.flush()
                if qty_to_set != 0:
                    record = StockRecord(
                        tenant_id=tenant_id,
                        inventory_id=inv.id,
                        record_type='IN' if qty_to_set > 0 else 'OUT',
                        change_quantity=qty_to_set,
                        balance_quantity=inv.quantity,
                        remark='Excel 批量导入' + (f' ({import_mode})' if import_mode != 'all' else ''),
                        operator_name=operator_name
                    )
                    db.session.add(record)
            else:
                for inv in invs:
                    # 仅在模式包含 data 时更新数量和成本
                    if import_mode in ['all', 'only_data']:
                        # 覆盖当前库存数量，而不是累加
                        diff_quantity = item['quantity'] - inv.quantity
                        if diff_quantity != 0:
                            inv.quantity = item['quantity']
                            record = StockRecord(
                                tenant_id=tenant_id,
                                inventory_id=inv.id,
                                record_type='IN' if diff_quantity > 0 else 'OUT',
                                change_quantity=diff_quantity,
                                balance_quantity=inv.quantity,
                                remark='Excel 批量导入 (覆盖同步)',
                                operator_name=operator_name
                            )
                            db.session.add(record)
                    
                    # 仅在模式包含 image 时更新图片
                    if import_mode in ['all', 'only_image'] and item.get('image_url'):
                        inv.image_url = item.get('image_url')
                    
                    # 仅在模式包含 data 时更新平均成本
                    if import_mode in ['all', 'only_data'] and item.get('avg_cost'):
                        inv.avg_cost = item.get('avg_cost')
                    
                    # 总是尝试更新系列信息（如果 Excel 中有的话）
                    if item.get('series'):
                        inv.series = item.get('series')
                db.session.flush()
            count += 1
            
        db.session.commit()
        return count, new_models

    @classmethod
    def export_inventory(cls, tenant_id, search='', status='', series=''):
        query = Inventory.query.filter_by(tenant_id=tenant_id)
        if search:
            query = query.filter(db.or_(
                Inventory.model.ilike(f'%{search}%'),
                Inventory.spec.ilike(f'%{search}%')
            ))
        if status:
            if status == 'NORMAL':
                query = query.filter(Inventory.quantity > 5)
            elif status == 'LOW':
                query = query.filter(Inventory.quantity > 0, Inventory.quantity <= 5)
            elif status == 'OUT':
                query = query.filter(Inventory.quantity <= 0)
        if series:
            query = query.filter(Inventory.series == series)
            
        items = query.order_by(Inventory.model.asc(), Inventory.spec.asc()).all()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "库存清单"
        
        # 使用更易读的、也支持标准导入的列头
        headers = ['图片', '型号', '规格', '数量', '平均成本', '系列']
        ws.append(headers)
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        
        base_upload = current_app.config.get('UPLOAD_FOLDER', 'uploads')
        
        for idx, item in enumerate(items, start=2):
            ws.row_dimensions[idx].height = 60
            
            ws.cell(row=idx, column=2, value=item.model).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=3, value=item.spec).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=4, value=float(item.quantity)).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=5, value=float(item.avg_cost)).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=idx, column=6, value=item.series).alignment = Alignment(horizontal='center', vertical='center')
            
            if item.image_url:
                filename = item.image_url.split('/')[-1]
                img_path = os.path.join(base_upload, 'inventory', filename)
                if os.path.exists(img_path):
                    try:
                        img = OpenpyxlImage(img_path)
                        img.width = 75
                        img.height = 75
                        ws.add_image(img, f'A{idx}')
                    except Exception as e:
                        pass
                        
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        filename = f"库存导出_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        return file_stream, filename
