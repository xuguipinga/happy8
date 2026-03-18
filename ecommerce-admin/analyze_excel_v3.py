import pandas as pd
import openpyxl

file_path = r'd:\WorkSpec\Project\Ecommerce-admin\ecommerce-admin\excel-model\克罗心最新表格.xlsx'

# Check pandas view
df = pd.read_excel(file_path, header=None)
print("Pandas columns:", df.columns.tolist())
print("First 5 rows:")
print(df.head(5))

# Check openpyxl view
wb = openpyxl.load_workbook(file_path)
ws = wb.active
print(f"Sheet: {ws.title}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")

# Check images
print(f"Number of images: {len(ws._images)}")
if len(ws._images) > 0:
    img = ws._images[0]
    print(f"First image anchor: row={img.anchor._from.row}, col={img.anchor._from.col}")
